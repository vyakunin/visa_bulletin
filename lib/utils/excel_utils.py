"""Reusable utilities for reading Excel files with openpyxl.

Provides common patterns for:
- Reading headers
- Streaming rows as dicts (memory efficient)
- Reading specific rows
- Counting rows

All functions handle proper workbook closing and error handling.
"""

import logging
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def read_excel_headers(filepath: Path, read_only: bool = True) -> list[str]:
    """
    Read column headers from first row of Excel file.
    
    Args:
        filepath: Path to Excel file (.xlsx or .xls)
        read_only: If True, use read_only mode (faster, less memory)
    
    Returns:
        List of header strings (empty string for None values)
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not Excel format
    """
    if filepath.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"Not an Excel file: {filepath.suffix}")

    wb = load_workbook(filepath, read_only=read_only)
    try:
        ws = wb.active
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else '')
        return headers
    finally:
        wb.close()


def _count_excel_rows(filepath: Path, read_only: bool = True) -> int:
    """
    Count data rows in Excel file (excluding header row).
    
    Args:
        filepath: Path to Excel file
        read_only: If True, use read_only mode (faster, less memory)
    
    Returns:
        Number of data rows (0 if file is empty or only has header)
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not Excel format
    """
    if filepath.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"Not an Excel file: {filepath.suffix}")

    wb = load_workbook(filepath, read_only=read_only)
    try:
        ws = wb.active
        # max_row includes header, so subtract 1 for data rows
        row_count = max(0, ws.max_row - 1) if ws.max_row else 0
        return row_count
    finally:
        wb.close()


def read_excel_streaming(
    filepath: Path,
    start_row: int = 2,
    read_only: bool = True,
    data_only: bool = True
) -> Iterator[dict]:
    """
    Stream Excel rows as dictionaries, starting from specified row.
    
    Args:
        filepath: Path to Excel file
        start_row: Row number to start from (1-indexed, default 2 to skip header)
        read_only: If True, use read_only mode (faster, less memory)
        data_only: If True, read calculated values instead of formulas
    
    Yields:
        Dictionary with column headers as keys and row values as values
    
    Example:
        >>> for record in read_excel_streaming(filepath):
        ...     print(record['CASE_NUMBER'])
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not Excel format
    """
    if filepath.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"Not an Excel file: {filepath.suffix}")

    wb = load_workbook(filepath, read_only=read_only, data_only=data_only)
    try:
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else '')

        # Stream rows starting from start_row
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            record = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    # Convert value to string, handle None
                    record[headers[i]] = str(value) if value is not None else ''
            yield record
    finally:
        wb.close()


def read_excel_row(
    filepath: Path,
    row_number: int,
    read_only: bool = True,
    data_only: bool = True
) -> dict | None:
    """
    Read a specific row from Excel file as dictionary.
    
    Wrapper around read_excel_rows() for convenience.
    
    Args:
        filepath: Path to Excel file
        row_number: Row number to read (1-indexed, 1 = header row)
        read_only: If True, use read_only mode (faster, less memory)
        data_only: If True, read calculated values instead of formulas
    
    Returns:
        Dictionary with column headers as keys and row values as values,
        or None if row doesn't exist
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not Excel format
    """
    rows = read_excel_rows(filepath, [row_number], read_only=read_only, data_only=data_only)
    return rows[0] if rows else None


def read_excel_rows(
    filepath: Path,
    row_numbers: list[int],
    read_only: bool = True,
    data_only: bool = True
) -> list[dict]:
    """
    Read multiple specific rows from Excel file as dictionaries.
    
    Optimized for large files: uses sequential iteration instead of random access
    for better performance with read_only mode.
    
    Args:
        filepath: Path to Excel file
        row_numbers: List of row numbers to read (1-indexed, 1 = header row)
        read_only: If True, use read_only mode (faster, less memory)
        data_only: If True, read calculated values instead of formulas
    
    Returns:
        List of dictionaries (one per row), in the order of row_numbers (not sorted)
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not Excel format
    """
    if filepath.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"Not an Excel file: {filepath.suffix}")

    if not row_numbers:
        return []

    # Filter and validate row numbers
    valid_rows = sorted(set(r for r in row_numbers if r > 0))
    if not valid_rows:
        return []

    wb = load_workbook(filepath, read_only=read_only, data_only=data_only)
    try:
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else '')

        # Check max_row to avoid iterating beyond file
        max_valid_row = min(valid_rows[-1], ws.max_row) if ws.max_row else 0
        if max_valid_row < valid_rows[0]:
            return []

        # Use sequential iteration for better performance with read_only mode
        # This is faster than random access (ws[row_num]) for large files
        needed_rows = set(valid_rows)
        records_dict = {}  # Store by row number

        # Iterate sequentially from first needed row to last
        min_row = valid_rows[0]
        max_row = max_valid_row

        # Track current row number as we iterate
        current_row_num = min_row
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            if current_row_num in needed_rows:
                record = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = str(value) if value is not None else ''
                records_dict[current_row_num] = record

                # Early exit if we've collected all needed rows
                if len(records_dict) >= len(needed_rows):
                    break

            current_row_num += 1

        # Return in original row_numbers order (not sorted)
        result = []
        for row_num in row_numbers:
            if row_num in records_dict:
                result.append(records_dict[row_num])

        return result
    finally:
        wb.close()


def get_excel_info(filepath: Path) -> dict:
    """
    Get basic information about Excel file (headers, row count, etc.).
    
    Args:
        filepath: Path to Excel file
    
    Returns:
        Dictionary with keys:
        - 'headers': List of column headers
        - 'row_count': Number of data rows (excluding header)
        - 'max_row': Maximum row number (including header)
        - 'filename': Filename
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not Excel format
    """
    if filepath.suffix.lower() not in ['.xlsx', '.xls']:
        raise ValueError(f"Not an Excel file: {filepath.suffix}")

    wb = load_workbook(filepath, read_only=True)
    try:
        ws = wb.active

        # Get headers
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else '')

        # Get row count
        row_count = max(0, ws.max_row - 1) if ws.max_row else 0

        return {
            'headers': headers,
            'row_count': row_count,
            'max_row': ws.max_row or 0,
            'filename': filepath.name,
            'filepath': str(filepath),
        }
    finally:
        wb.close()

