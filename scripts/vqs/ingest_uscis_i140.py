#!/usr/bin/env python3
"""
Ingest USCIS I-140 receipts into raw_facts_ledger for VQS.

Usage:
  # Insert stub data (for MVP when no USCIS file available)
  bazel run //scripts/vqs:ingest_uscis_i140 -- --stub

  # Ingest from USCIS XLSX file (e.g. i140_rec_by_class_country_fy2024_q3.xlsx)
  bazel run //scripts/vqs:ingest_uscis_i140 -- --file path/to/file.xlsx [--publication-date YYYY-MM-DD]

When to use:
- After downloading USCIS "I-140 Receipts by Category and Country" quarterly data
- Use --stub to populate ledger for testing the solver without real files
"""

import argparse
import logging
import os
from datetime import date, timedelta
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "django_config.settings")
import django

django.setup()

from django.db.utils import IntegrityError

from django_config.logging_config import setup_logging
from lib.utils.logging_utils import ScriptLogger
from models.enums.country import Country
from models.raw_facts import RawFactsLedger, RawFactSource

setup_logging(debug=False)
logger = logging.getLogger(__name__)
script_logger = ScriptLogger(__file__)

# USCIS column name patterns (format may vary by year)
PREFERENCE_MAP = {
    "EB-1": "1st",
    "EB-2": "2nd",
    "EB-3": "3rd",
    "EB-4": "4th",
    "EB-5": "5th",
    "EB1": "1st",
    "EB2": "2nd",
    "EB3": "3rd",
    "EB4": "4th",
    "EB5": "5th",
    "First Preference": "1st",
    "Second Preference": "2nd",
    "Third Preference": "3rd",
    "Fourth Preference": "4th",
    "Fifth Preference": "5th",
    "1st": "1st",
    "2nd": "2nd",
    "3rd": "3rd",
    "4th": "4th",
    "5th": "5th",
}
# Map USCIS string to Country enum value (int). First step in all country handling is to convert to enum.
COUNTRY_TO_ENUM = {
    "All Other Countries": Country.ALL.value,
    "All Other": Country.ALL.value,
    "Other Countries": Country.ALL.value,
    "Other": Country.ALL.value,
    "China": Country.CHINA.value,
    "China (mainland born)": Country.CHINA.value,
    "India": Country.INDIA.value,
    "Mexico": Country.MEXICO.value,
    "Philippines": Country.PHILIPPINES.value,
    "El Salvador Guatemala Honduras": Country.EL_SALVADOR_GUATEMALA_HONDURAS.value,
    "El Salvador/Guatemala/Honduras": Country.EL_SALVADOR_GUATEMALA_HONDURAS.value,
}


def _quarter_dates(fiscal_year: int, quarter: int) -> tuple[date, date]:
    """Return (start_date, end_date) for FY quarter. FY Q1 = Oct-Dec."""
    start_month = (quarter - 1) * 3 + 10
    if start_month > 12:
        start_month -= 12
        start_year = fiscal_year
    else:
        start_year = fiscal_year - 1
    start_date = date(start_year, start_month, 1)
    end_date = start_date + timedelta(days=92)
    end_date = end_date.replace(day=1) - timedelta(days=1)
    return start_date, end_date


def insert_stub_data(publication_date: date | None = None) -> int:
    """Insert stub I-140 receipt rows for MVP testing. Returns count inserted."""
    pub = publication_date or date.today()
    count = 0
    # FY2024 Q2, Q3: a few (preference, country) pairs
    for fy, q in [(2024, 2), (2024, 3)]:
        start_d, end_d = _quarter_dates(fy, q)
        for category in ["2nd", "3rd"]:
            for country_enum in [Country.INDIA, Country.CHINA, Country.ALL]:
                dimensions = {"country": country_enum.value, "category": category}
                value = 1000 + (abs(hash((fy, q, category, country_enum.value))) % 4000)
                try:
                    RawFactsLedger.objects.create(
                        source=RawFactSource.USCIS_I140,
                        metric="i140_receipts",
                        dimensions=dimensions,
                        value=value,
                        reference_period_start=start_d,
                        reference_period_end=end_d,
                        publication_date=pub,
                    )
                    count += 1
                except IntegrityError:
                    logger.debug("Skip duplicate stub row %s %s %s", fy, q, dimensions)
    logger.info("Inserted %d stub I-140 receipt rows", count)
    return count


# Sheet name to country (e.g. "India FY25" -> India). Only VQS-relevant countries to avoid duplicate (source, metric, dimensions, period).
SHEET_COUNTRY_PATTERNS = [
    ("All Countries", Country.ALL.value),
    ("India", Country.INDIA.value),
    ("China", Country.CHINA.value),
    ("Mexico", Country.MEXICO.value),
    ("Philippines", Country.PHILIPPINES.value),
    ("El Salvador", Country.EL_SALVADOR_GUATEMALA_HONDURAS.value),
]


def _country_from_sheet_name(sheet_title: str) -> int | None:
    """Return country enum value from sheet title, or None if not mapped."""
    title = sheet_title or ""
    for pattern, enum_val in SHEET_COUNTRY_PATTERNS:
        if pattern.lower() in title.lower():
            return enum_val
    return None


def _parse_xlsx(path: Path) -> list[tuple[int, int, str, int, int]]:
    """Parse USCIS I-140 XLSX (by-preference, by-country sheets). Returns list of (fy, q, category, country_enum_value, count)."""
    try:
        import openpyxl
    except ImportError as e:
        logger.error("openpyxl required for XLSX: %s", e)
        return []
    rows = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        country_val = _country_from_sheet_name(sheet.title)
        if country_val is None:
            continue
        sheet_rows = list(sheet.iter_rows(values_only=True))
        if len(sheet_rows) < 5:
            continue
        # Find header row dynamically — some files have it at row 3, others at row 4
        header_idx = None
        for idx in range(min(10, len(sheet_rows))):
            row_cells = sheet_rows[idx]
            year_count = sum(
                1 for c in row_cells if isinstance(c, int) and 2010 <= c <= 2030
            )
            if year_count >= 3:
                header_idx = idx
                break
        if header_idx is None:
            continue
        header_row = sheet_rows[header_idx]
        fy_columns: list[tuple[int, int]] = []  # (col_index, year)
        for i, cell in enumerate(header_row):
            if isinstance(cell, int) and 2010 <= cell <= 2030:
                fy_columns.append((i, cell))
        if not fy_columns:
            continue
        current_cat: str | None = None
        for row in sheet_rows[header_idx + 1:]:
            if not row or len(row) <= fy_columns[0][0]:
                continue
            label = (row[0] or "").strip() if isinstance(row[0], str) else ""
            if "First Preference (EB1)" in label or label == "First Preference (EB1)":
                current_cat = "1st"
                continue
            if "Second Preference (EB2)" in label or label == "Second Preference (EB2)":
                current_cat = "2nd"
                continue
            if "Third Preference (EB3)" in label or label == "Third Preference (EB3)":
                current_cat = "3rd"
                continue
            if current_cat and label == "Total Petitions":
                for col_idx, fy in fy_columns:
                    if col_idx < len(row):
                        cell = row[col_idx]
                        if isinstance(cell, (int, float)) and cell >= 0:
                            cnt = int(cell)
                            if cnt > 0:
                                # Spread annual total across 4 fiscal quarters
                                # for better temporal distribution in demand model
                                per_q = cnt // 4
                                leftover = cnt - per_q * 4
                                for q in range(1, 5):
                                    q_count = per_q + (1 if q <= leftover else 0)
                                    if q_count > 0:
                                        rows.append(
                                            (fy, q, current_cat, country_val, q_count)
                                        )
                continue
    wb.close()
    return rows


def ingest_file(file_path: Path, publication_date: date | None = None) -> int:
    """Parse XLSX and insert into ledger. Returns count inserted.

    If --publication-date is given, all rows use that date (legacy behavior).
    Otherwise, each row's publication_date is set to reference_period_end + 90 days
    (approximate USCIS publication delay), enabling correct backtesting.
    """
    use_historical = publication_date is None
    parsed = _parse_xlsx(file_path)
    if not parsed:
        logger.warning("No rows parsed from %s", file_path)
        return 0
    count = 0
    for fy, q, category, country_value, receipt_count in parsed:
        start_d, end_d = _quarter_dates(fy, q)
        # Historical pub date: ~90 days after quarter end (USCIS publication delay)
        pub = end_d + timedelta(days=90) if use_historical else publication_date
        dimensions = {"country": country_value, "category": category}
        try:
            RawFactsLedger.objects.update_or_create(
                source=RawFactSource.USCIS_I140,
                metric="i140_receipts",
                dimensions=dimensions,
                reference_period_start=start_d,
                reference_period_end=end_d,
                defaults={
                    "value": receipt_count,
                    "publication_date": pub,
                },
            )
            count += 1
        except IntegrityError:
            logger.debug("Skip duplicate %s Q%s %s", fy, q, dimensions)
    logger.info(
        "Inserted %d I-140 receipt rows from %s (historical_pub=%s)",
        count,
        file_path,
        use_historical,
    )
    return count


def _inspect_xlsx(path: Path, max_rows: int = 30) -> None:
    """Print raw rows from first sheet for debugging."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for si, sheet in enumerate(wb.worksheets):
        logger.info("Sheet %d: %s", si, sheet.title)
        for ri, row in enumerate(sheet.iter_rows(values_only=True)):
            if ri >= max_rows:
                break
            logger.info("  row %d: %s", ri, row)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ingest USCIS I-140 receipts into raw_facts_ledger"
    )
    parser.add_argument(
        "--stub", action="store_true", help="Insert stub data for MVP testing"
    )
    parser.add_argument("--file", type=Path, help="Path to USCIS I-140 XLSX file")
    parser.add_argument(
        "--publication-date",
        type=str,
        help="Publication date YYYY-MM-DD (default: today)",
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Print raw rows from XLSX (with --file) and exit",
    )
    args = parser.parse_args()

    script_logger.log_call(
        args={"stub": args.stub, "file": str(args.file) if args.file else None},
        context="VQS: Ingest USCIS I-140 into raw_facts_ledger",
    )

    pub_date = None
    if args.publication_date:
        pub_date = date.fromisoformat(args.publication_date)

    if args.stub:
        insert_stub_data(publication_date=pub_date)
    elif args.file:
        if not args.file.exists():
            logger.error("File not found: %s", args.file)
            raise SystemExit(1)
        if args.inspect:
            _inspect_xlsx(args.file)
        else:
            ingest_file(args.file, publication_date=pub_date)
    else:
        parser.error("Specify --stub or --file PATH")


if __name__ == "__main__":
    main()
