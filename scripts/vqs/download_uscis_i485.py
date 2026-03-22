#!/usr/bin/env python3
"""
Download and ingest USCIS EB Pending I-485 Inventory XLSX reports.

USCIS publishes monthly "Pending Applications for Employment-Based Preference
Categories" reports as XLSX files. Each snapshot provides pending I-485 counts
by preference category (EB-1, EB-2, EB-3), country of chargeability (India,
China, Philippines, Mexico, All), and priority date month/year.

Usage:
  # Download all available files and ingest
  bazel run //scripts/vqs:download_uscis_i485 -- --output-dir /tmp/i485_data

  # Download only (skip ingest)
  bazel run //scripts/vqs:download_uscis_i485 -- --output-dir /tmp/i485_data --download-only

  # List URLs that will be tried (no download)
  bazel run //scripts/vqs:download_uscis_i485 -- --list-urls

  # Ingest a previously downloaded file (skips download)
  bazel run //scripts/vqs:download_uscis_i485 -- --file /tmp/i485_data/eb_inventory_october_2025.xlsx --publication-date 2025-10-02

  # Inspect XLSX structure without ingesting (debug mode)
  bazel run //scripts/vqs:download_uscis_i485 -- --file /tmp/i485_data/eb_inventory_october_2025.xlsx --inspect

Data notes:
- Available monthly from approximately July 2022 onwards
- Snapshot date is the "as of" date in the title (first week of each month)
- Publication date (when USCIS releases it) is usually 4-6 weeks after snapshot
- Stored as metric 'i485_pending_inventory_monthly' in RawFactsLedger
- Source: https://www.uscis.gov/tools/reports-and-studies/immigration-and-citizenship-data
"""

import argparse
import calendar
import logging
import os
from datetime import date
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "django_config.settings")
import django

django.setup()

from django_config.logging_config import setup_logging
from lib.utils.logging_utils import ScriptLogger
from models.enums.country import Country

setup_logging(debug=False)
logger = logging.getLogger(__name__)
script_logger = ScriptLogger(__file__)

# USCIS uses two base paths inconsistently. Try both for each file.
USCIS_DATA_BASE = "https://www.uscis.gov/sites/default/files/document/data"
USCIS_REPORTS_BASE = "https://www.uscis.gov/sites/default/files/document/reports"

MONTH_NAMES = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]

# Confirmed URLs from USCIS data page (as of March 2026).
# Tuple: (filename_stem, snapshot_date, publication_date)
# snapshot_date = the "as of" date in the title; pub_date = USCIS release date
KNOWN_FILES: list[tuple[str, date, date]] = [
    ("eb_inventory_october_2025",  date(2025, 10, 2),  date(2025, 11, 12)),
    ("eb_inventory_september_2025", date(2025, 9, 3),  date(2025, 10, 14)),
    ("eb_inventory_august_2025",   date(2025, 8, 5),   date(2025, 10, 14)),
]


def _generate_urls() -> list[tuple[str, str, date, date]]:
    """Generate (url, description, snapshot_date, publication_date) tuples to try.

    Combines confirmed known files with speculative monthly URLs back to 2022.
    Tries both /document/data/ and /document/reports/ paths since USCIS is
    inconsistent about which path they use.
    """
    seen: set[str] = set()
    results: list[tuple[str, str, date, date]] = []

    # Add confirmed known files first (both URL base paths)
    for stem, snap_date, pub_date in KNOWN_FILES:
        filename = f"{stem}.xlsx"
        for base in [USCIS_DATA_BASE, USCIS_REPORTS_BASE]:
            url = f"{base}/{filename}"
            if url not in seen:
                seen.add(url)
                results.append((url, f"Confirmed: {stem}", snap_date, pub_date))

    # Speculative: generate monthly files from July 2022 to present
    # (USCIS started publishing these reports around mid-2022)
    start_year, start_month = 2022, 7
    today = date.today()
    year, month = start_year, start_month
    while (year, month) <= (today.year, today.month):
        month_name = MONTH_NAMES[month - 1]
        stem = f"eb_inventory_{month_name}_{year}"
        snap_date = date(year, month, 1)
        # Publication typically 4-6 weeks after snapshot
        pub_year_actual = year if month + 2 <= 12 else year + 1
        pub_date = date(pub_year_actual, ((month + 1) % 12) + 1, 15)

        filename = f"{stem}.xlsx"
        for base in [USCIS_DATA_BASE, USCIS_REPORTS_BASE]:
            url = f"{base}/{filename}"
            if url not in seen:
                seen.add(url)
                results.append((url, f"Speculative {month_name.capitalize()} {year}", snap_date, pub_date))

        month += 1
        if month > 12:
            month = 1
            year += 1

    return results


def _download_file(url: str, output_dir: Path) -> Path | None:
    """Download a file from URL. Returns local path or None if not found."""
    import requests

    filename = url.split("/")[-1]
    local_path = output_dir / filename

    if local_path.exists():
        logger.info("Already downloaded: %s", filename)
        return local_path

    try:
        resp = requests.get(url, timeout=30, allow_redirects=True)
        if resp.status_code == 200:
            content_type = resp.headers.get("content-type", "")
            if "html" in content_type.lower():
                # USCIS serves a 404 HTML page for missing files (not 404 status)
                logger.debug("Got HTML for %s — likely not found", url)
                return None
            if len(resp.content) < 10_000:
                # USCIS XLSX files are typically 50KB+; too small = redirect/error page
                logger.debug("Too small (%d bytes) for %s", len(resp.content), url)
                return None
            local_path.write_bytes(resp.content)
            logger.info("Downloaded: %s (%d KB)", filename, len(resp.content) // 1024)
            return local_path
        else:
            logger.debug("HTTP %d for %s", resp.status_code, url)
            return None
    except Exception as e:
        logger.debug("Failed to download %s: %s", url, e)
        return None


# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

# Maps USCIS preference label → VQS visa_class string
_PREF_MAP: dict[str, str] = {
    "1st": "1st", "first": "1st", "eb-1": "1st", "eb1": "1st",
    "1st preference": "1st", "employment first": "1st",
    "2nd": "2nd", "second": "2nd", "eb-2": "2nd", "eb2": "2nd",
    "2nd preference": "2nd", "employment second": "2nd",
    "3rd": "3rd", "third": "3rd", "eb-3": "3rd", "eb3": "3rd",
    "3rd preference": "3rd", "employment third": "3rd",
    "other workers": "3rd",  # EW shares 3rd quota; treat as 3rd for queue depth
    "ew": "3rd",
}

# Maps USCIS country label → Country enum value.
# Keys must be exact lowercase strings as they appear in USCIS cells or sheet titles.
_COUNTRY_MAP: dict[str, int] = {
    "india": Country.INDIA.value,
    "china": Country.CHINA.value,
    "china (mainland born)": Country.CHINA.value,
    "mainland china": Country.CHINA.value,
    "philippines": Country.PHILIPPINES.value,
    "mexico": Country.MEXICO.value,
    "all chargeability": Country.ALL.value,
    "all chargeability areas": Country.ALL.value,
    "all other": Country.ALL.value,
    "rest of the world": Country.ALL.value,
    "rest of world": Country.ALL.value,
    "row": Country.ALL.value,
    "all": Country.ALL.value,
}


def _normalize_pref(label: str) -> str | None:
    """Map USCIS preference label to VQS visa_class. Handles full USCIS labels via substring search."""
    key = label.strip().lower()
    if key in _PREF_MAP:
        return _PREF_MAP[key]
    # USCIS full labels: 'Employment-Based 1st Preference Category (EB1)', etc.
    # Use substring matching on the embedded short codes
    if "1st" in key or "eb-1" in key or "eb1" in key or "first preference" in key:
        return "1st"
    if "2nd" in key or "eb-2" in key or "eb2" in key or "second preference" in key:
        return "2nd"
    if "3rd" in key or "eb-3" in key or "eb3" in key or "third preference" in key:
        return "3rd"
    if "other workers" in key or " ew" in key or "(ew" in key:
        return "3rd"
    return None


def _normalize_country(label: str) -> int | None:
    """Map USCIS country string to Country enum value. Tries exact match first, then prefix match."""
    key = label.strip().lower()
    if key in _COUNTRY_MAP:
        return _COUNTRY_MAP[key]
    # Fallback: check if any known label is contained in the cell value
    # (e.g. "China (mainland born)" contains "china")
    for known, val in _COUNTRY_MAP.items():
        if len(known) >= 4 and known in key:
            return val
    return None


def _inspect_xlsx(path: Path, max_rows: int = 20) -> None:
    """Print raw cell values for all sheets (debug mode)."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required. Add requirement('openpyxl') to deps.")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for si, sheet in enumerate(wb.worksheets):
        print(f"\n=== Sheet {si}: '{sheet.title}' ===")
        for ri, row in enumerate(sheet.iter_rows(values_only=True)):
            if ri >= max_rows:
                print(f"  ... (truncated at {max_rows} rows)")
                break
            non_none = [c for c in row if c is not None]
            if non_none:
                print(f"  row {ri:3d}: {list(row[:20])}")
    wb.close()


def _parse_xlsx(path: Path) -> list[tuple[str, int, int, int, int]]:
    """Parse USCIS EB I-485 inventory XLSX.

    Returns list of (visa_class, country_enum, priority_year, priority_month, count).

    XLSX format (confirmed from 2024-2025 files):
    - Multiple sheets (by country: 'Rest of the World', 'China', 'India (EB1...)', 'India (EB2 EB3)')
    - Each sheet has a flat table starting at row 3 (0-indexed) with header:
      ['Country Of Chargeability', 'Preference Category', 'Visa Status',
       'Priority Date Month',
       'Priority Date Year - Prior Years',
       'Priority Date Year - 2016', ..., 'Priority Date Year - 2025']
    - Data rows: one row per (country, preference, visa_status, priority_month) combination
    - Values: integer | 'D' (suppressed, <10; we use 5 as estimate) | '-' (zero)
    - Visa Status is either 'Available' or 'Awaiting Availability' — we include both
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required")
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    results: list[tuple[str, int, int, int, int]] = []

    for sheet in wb.worksheets:
        sheet_rows = list(sheet.iter_rows(values_only=True))
        logger.debug("Parsing sheet '%s' (%d rows)", sheet.title, len(sheet_rows))
        parsed = _parse_sheet(sheet.title, sheet_rows)
        logger.debug("  → %d records extracted", len(parsed))
        results.extend(parsed)

    wb.close()

    if not results:
        logger.warning(
            "No rows extracted from %s. Run with --inspect to see the raw structure.",
            path.name,
        )
    return results


def _parse_sheet(
    sheet_title: str,
    rows: list[tuple],
) -> list[tuple[str, int, int, int, int]]:
    """Parse one data sheet. Returns (visa_class, country_enum, priority_year, priority_month, count).

    Handles the confirmed USCIS flat-table format with columns:
    Country Of Chargeability | Preference Category | Visa Status |
    Priority Date Month | Priority Date Year - Prior Years | Priority Date Year - YYYY | ...
    """
    import re

    results: list[tuple[str, int, int, int, int]] = []
    if not rows:
        return results

    # Find the header row: must contain 'Country Of Chargeability'
    header_idx: int | None = None
    for i, row in enumerate(rows):
        if row:
            for c in row:
                if c and isinstance(c, str) and "country of chargeability" in c.lower():
                    header_idx = i
                    break
        if header_idx is not None:
            break

    if header_idx is None:
        logger.debug(
            "  Sheet '%s': no header row found. First row sample: %s",
            sheet_title,
            (rows[0][:5] if rows else "empty"),
        )
        return results  # Not a data sheet (e.g. 'How to Read This Report')

    header = rows[header_idx]

    # Map column indices
    country_col = preference_col = month_col = None
    year_columns: list[tuple[int, int | None]] = []  # (col_idx, year_int or None for 'Prior Years')

    for ci, cell in enumerate(header):
        if not cell or not isinstance(cell, str):
            continue
        s = cell.strip()
        sl = s.lower()
        if "country of chargeability" in sl:
            country_col = ci
        elif "preference category" in sl:
            preference_col = ci
        elif "priority date month" in sl:
            month_col = ci
        elif "priority date year" in sl:
                if "prior years" in sl:
                    year_columns.append((ci, None))  # 'Prior Years' — skip in output
                else:
                    m = re.search(r"(\d{4})", s)
                    if m:
                        year_columns.append((ci, int(m.group(1))))

    if month_col is None or not year_columns:
        logger.debug("  Sheet '%s': skipping (month_col=%s year_cols=%d)", sheet_title, month_col, len(year_columns))
        return results

    logger.debug(
        "  Sheet '%s': %d data rows, %d year columns",
        sheet_title, len(rows) - header_idx - 1, len(year_columns),
    )

    for row in rows[header_idx + 1:]:
        if not row or len(row) <= month_col:
            continue

        # Priority date month
        month_raw = row[month_col]
        if month_raw is None:
            continue
        priority_month = _parse_month_label(str(month_raw))
        if priority_month is None:
            continue

        # Country
        country: int | None = None
        if country_col is not None and row[country_col]:
            country = _normalize_country(str(row[country_col]))
        if country is None:
            country = _country_from_sheet_title(sheet_title)
        if country is None:
            continue

        # Preference category
        pref: str | None = None
        if preference_col is not None and row[preference_col]:
            pref = _normalize_pref(str(row[preference_col]))
        if pref is None:
            continue
        # Extract counts for each explicit year column (skip "Prior Years")
        for col_idx, year in year_columns:
            if year is None:
                continue  # Skip pre-range aggregate column
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if cell is None or cell == "-":
                continue
            if cell == "D":
                # Suppressed (< 10 applications). Use 5 as midpoint estimate.
                count = 5
            elif isinstance(cell, (int, float)):
                count = int(cell)
            else:
                try:
                    count = int(str(cell).replace(",", ""))
                except ValueError:
                    continue

            if count > 0:
                results.append((pref, country, year, priority_month, count))

    return results


def _country_from_sheet_title(title: str) -> int | None:
    """Extract country from sheet title like 'India (EB2 EB3)' or 'Rest of the World'."""
    tl = title.lower()
    for label, val in _COUNTRY_MAP.items():
        if label in tl:
            return val
    return None


def _parse_month_label(label: str) -> int | None:
    """Return 1-12 for month labels like 'Jan', 'January', '1', etc. or None."""
    label = label.strip()
    if not label:
        return None
    # Numeric month
    try:
        val = int(label)
        if 1 <= val <= 12:
            return val
    except ValueError:
        pass
    # 3-letter abbreviated month
    abbrev = label[:3].lower()
    month_abbrevs = ["jan", "feb", "mar", "apr", "may", "jun",
                     "jul", "aug", "sep", "oct", "nov", "dec"]
    if abbrev in month_abbrevs:
        return month_abbrevs.index(abbrev) + 1
    # Full month name
    full_lower = label.lower()
    for i, m in enumerate(MONTH_NAMES):
        if full_lower == m:
            return i + 1
    return None


# ---------------------------------------------------------------------------
# Ingestion
# ---------------------------------------------------------------------------

def ingest_file(path: Path, publication_date: date) -> int:
    """Parse XLSX and upsert facts into RawFactsLedger. Returns count inserted/updated."""
    from django.db import transaction

    from models.raw_facts import RawFactsLedger, RawFactSource

    parsed = _parse_xlsx(path)
    if not parsed:
        logger.warning("No records parsed from %s", path.name)
        return 0

    # Aggregate by (visa_class, country, year, month) — in case of duplicates from multi-sheet
    agg: dict[tuple[str, int, int, int], int] = {}
    for visa_class, country, year, month, count in parsed:
        key = (visa_class, country, year, month)
        agg[key] = agg.get(key, 0) + count

    to_create = []
    for (visa_class, country, year, month), total_count in agg.items():
        ref_start = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        ref_end = date(year, month, last_day)

        to_create.append(RawFactsLedger(
            source=RawFactSource.USCIS_I485_INVENTORY,
            metric="i485_pending_inventory_monthly",
            dimensions={"country": country, "visa_class": visa_class},
            value=total_count,
            reference_period_start=ref_start,
            reference_period_end=ref_end,
            publication_date=publication_date,
        ))

    with transaction.atomic():
        # Upsert: update value+publication_date if same (source, metric, dims, period) already exists.
        # This handles re-ingesting a revised file without duplicate-key errors.
        RawFactsLedger.objects.bulk_create(
            to_create,
            update_conflicts=True,
            unique_fields=["source", "metric", "dimensions", "reference_period_start", "reference_period_end"],
            update_fields=["value", "publication_date"],
            batch_size=1000,
        )

    logger.info(
        "Ingested %d I-485 inventory facts from %s (pub_date=%s)",
        len(to_create), path.name, publication_date,
    )
    return len(to_create)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Download and ingest USCIS EB pending I-485 inventory data"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("/tmp/uscis_i485_data"),
        help="Directory to save downloaded XLSX files (default: /tmp/uscis_i485_data)",
    )
    parser.add_argument(
        "--download-only",
        action="store_true",
        help="Download files but don't ingest into database",
    )
    parser.add_argument(
        "--list-urls",
        action="store_true",
        help="Print all URLs that will be attempted (no download)",
    )
    parser.add_argument(
        "--file",
        type=Path,
        help="Ingest (or inspect) a specific local XLSX file instead of downloading",
    )
    parser.add_argument(
        "--publication-date",
        type=str,
        help=(
            "Publication date YYYY-MM-DD for --file mode. "
            "Defaults to the snapshot date embedded in the filename."
        ),
    )
    parser.add_argument(
        "--inspect",
        action="store_true",
        help="Print raw XLSX cell values (with --file) and exit without ingesting",
    )
    args = parser.parse_args()

    script_logger.log_call(
        args={
            "file": str(args.file) if args.file else None,
            "output_dir": str(args.output_dir),
            "download_only": args.download_only,
            "publication_date": args.publication_date,
        },
        context="VQS: Download/ingest USCIS I-485 inventory",
    )

    # -- Single-file mode --
    if args.file:
        if not args.file.exists():
            logger.error("File not found: %s", args.file)
            raise SystemExit(1)
        if args.inspect:
            _inspect_xlsx(args.file)
            return
        if not args.publication_date:
            # Try to extract snapshot date from filename (e.g. eb_inventory_october_2025.xlsx)
            pub_date = _extract_date_from_filename(args.file.stem)
            if pub_date is None:
                logger.error(
                    "Cannot determine publication date from filename '%s'. "
                    "Pass --publication-date YYYY-MM-DD.",
                    args.file.stem,
                )
                raise SystemExit(1)
            logger.info("Inferred publication date from filename: %s", pub_date)
        else:
            pub_date = date.fromisoformat(args.publication_date)
        ingest_file(args.file, pub_date)
        return

    # -- Download mode --
    urls = _generate_urls()

    if args.list_urls:
        for url, desc, snap_date, pub_date in urls:
            print(f"{desc:50s}  snap={snap_date}  pub={pub_date}  {url}")
        print(f"\nTotal URLs: {len(urls)}")
        return

    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Deduplicate by filename: if both /data/ and /reports/ succeed, keep first
    downloaded: list[tuple[Path, date, date]] = []  # (path, snapshot_date, pub_date)
    seen_filenames: set[str] = set()

    for url, desc, snap_date, pub_date in urls:
        filename = url.split("/")[-1]
        if filename in seen_filenames:
            continue
        path = _download_file(url, args.output_dir)
        if path:
            seen_filenames.add(filename)
            downloaded.append((path, snap_date, pub_date))

    logger.info("Downloaded %d files", len(downloaded))

    if not downloaded:
        logger.warning(
            "No files downloaded. USCIS may have changed URL patterns. "
            "Check https://www.uscis.gov/tools/reports-and-studies/immigration-and-citizenship-data "
            "and download manually, then run:\n"
            "  bazel run //scripts/vqs:download_uscis_i485 -- "
            "--file <path> --publication-date YYYY-MM-DD"
        )
        return

    if args.download_only:
        for path, snap_date, _ in downloaded:
            logger.info("  snap=%s  %s", snap_date, path)
        return

    total = 0
    for path, snap_date, pub_date in downloaded:
        logger.info("Ingesting %s (pub_date=%s)...", path.name, pub_date)
        try:
            count = ingest_file(path, publication_date=pub_date)
            total += count
        except Exception as e:
            logger.error("Failed to ingest %s: %s", path.name, e)

    logger.info("Total: %d I-485 inventory facts ingested from %d files", total, len(downloaded))

    # Report ledger coverage
    from models.raw_facts import RawFactsLedger, RawFactSource

    row_count = RawFactsLedger.objects.filter(
        source=RawFactSource.USCIS_I485_INVENTORY,
        metric="i485_pending_inventory_monthly",
    ).count()
    logger.info("Ledger now has %d I-485 inventory facts total", row_count)

    earliest = (
        RawFactsLedger.objects.filter(source=RawFactSource.USCIS_I485_INVENTORY)
        .order_by("publication_date")
        .values_list("publication_date", flat=True)
        .first()
    )
    latest = (
        RawFactsLedger.objects.filter(source=RawFactSource.USCIS_I485_INVENTORY)
        .order_by("-publication_date")
        .values_list("publication_date", flat=True)
        .first()
    )
    if earliest and latest:
        logger.info("I-485 inventory coverage: pub_date %s to %s", earliest, latest)


def _extract_date_from_filename(stem: str) -> date | None:
    """Extract snapshot month/year from filename like 'eb_inventory_october_2025'."""
    parts = stem.lower().split("_")
    # Pattern: ['eb', 'inventory', 'october', '2025']
    for i, part in enumerate(parts):
        if part in MONTH_NAMES and i + 1 < len(parts):
            try:
                year = int(parts[i + 1])
                month = MONTH_NAMES.index(part) + 1
                return date(year, month, 1)
            except (ValueError, IndexError):
                pass
    return None


if __name__ == "__main__":
    main()
