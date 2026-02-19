#!/usr/bin/env python3
"""
Inspect available columns and sample data in DOL source files.

Usage:
    bazel run //scripts/ingest:inspect_source_columns
    bazel run //scripts/ingest:inspect_source_columns -- --files data/salary/dol_data/PERM_Disclosure_Data_FY2020.xlsx
    bazel run //scripts/ingest:inspect_source_columns -- --files data/salary/dol_data/LCA_Disclosure_Data_FY2024_Q4.xlsx
    bazel run //scripts/ingest:inspect_source_columns -- --match-terms job,title --show-all-columns
    bazel run //scripts/ingest:inspect_source_columns -- --estimate-storage --avg-length 30 --varchar-size 100
    
    # Smoke test: validate parsing before ingestion
    bazel run //scripts/ingest:inspect_source_columns -- --validate-parsing
    bazel run //scripts/ingest:inspect_source_columns -- --validate-parsing --files data/salary/dol_data/*.xlsx

Output:
    - Total columns in each file
    - Matching columns (by keyword filters) with sample values
    - Total record counts per file
    - Optional storage impact estimates
    - Parsing validation results (with --validate-parsing flag)
"""

import argparse
import logging
import os
import sys
from pathlib import Path

# Setup Django early for validation mode
if '--validate-parsing' in sys.argv:
    if not os.environ.get('DJANGO_SETTINGS_MODULE'):
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'django_config.settings')
    import django
    django.setup()

from django_config.logging_config import setup_logging
from lib.utils.data_source_utils import get_file_stats
from lib.utils.excel_utils import (
    read_excel_headers,
    read_excel_rows,
)
from lib.utils.http_utils import get_workspace_dir
from lib.utils.logging_utils import ScriptLogger

DEFAULT_FILES = [
    "data/salary/dol_data/PERM_Disclosure_Data_FY2020.xlsx",
    "data/salary/dol_data/LCA_Disclosure_Data_FY2024_Q4.xlsx",
]
DEFAULT_MATCH_TERMS = ["name", "alien", "beneficiary", "employee", "worker"]
DEFAULT_EXCLUDE_TERMS = ["employer"]


def _resolve_files(files: list[str]) -> list[Path]:
    workspace_dir = get_workspace_dir()
    resolved = []
    for file_path in files:
        candidate = Path(file_path)
        if not candidate.is_absolute():
            candidate = workspace_dir / candidate
        resolved.append(candidate)
    return resolved


def _find_matching_columns(
    columns: list[str],
    match_terms: list[str],
    exclude_terms: list[str],
) -> list[str]:
    match_terms_lower = [term.lower() for term in match_terms]
    exclude_terms_lower = [term.lower() for term in exclude_terms]
    matches = []
    for column in columns:
        column_str = str(column).lower()
        if any(term in column_str for term in match_terms_lower) and not any(
            term in column_str for term in exclude_terms_lower
        ):
            matches.append(column)
    return matches


def _log_column_samples(
    logger: logging.Logger,
    file_path: Path,
    columns: list[str],
    sample_rows: int,
    match_terms: list[str],
    exclude_terms: list[str],
    show_all_columns: bool,
) -> None:
    logger.info("Checking %s...", file_path.name)
    logger.info("=" * 80)

    if show_all_columns:
        logger.info("Total columns: %s", len(columns))
        logger.info("All columns:")
        for i, column in enumerate(columns, 1):
            logger.info("%3d. %s", i, column)

    matching_columns = _find_matching_columns(columns, match_terms, exclude_terms)
    logger.info("=" * 80)
    logger.info("Matching columns:")
    if not matching_columns:
        logger.info("  (none found)")
        return

    sample_row_numbers = list(range(2, 2 + sample_rows))
    sample_data = read_excel_rows(file_path, sample_row_numbers)
    for column in matching_columns:
        logger.info("  ✓ %s", column)
        samples = [row.get(column, "") for row in sample_data if row.get(column, "")]
        if samples:
            logger.info("    Sample: %s", samples[:2])


def _log_storage_estimate(
    logger: logging.Logger,
    total_records: int,
    avg_length: int,
    varchar_size: int,
) -> None:
    bytes_per_name = avg_length + 4
    total_bytes = total_records * bytes_per_name
    total_mb = total_bytes / (1024 * 1024)
    total_gb = total_mb / 1024

    logger.info("")
    logger.info("=" * 80)
    logger.info("STORAGE IMPACT ESTIMATE:")
    logger.info("=" * 80)
    logger.info("Average value length: %s characters", avg_length)
    logger.info("Column type: VARCHAR(%s)", varchar_size)
    logger.info("Storage per record: ~%s bytes", bytes_per_name)
    logger.info("Total storage: %.1f MB (%.2f GB)", total_mb, total_gb)
    logger.info("")
    logger.info("For comparison:")
    logger.info("  Current SalaryRecord has ~20 columns")
    logger.info("  Adding 1 name column = ~5%% increase in table size")


def _validate_parsing(
    logger: logging.Logger,
    file_path: Path,
    sample_rows: int = 5,
) -> dict:
    """
    Smoke test: parse sample rows through plugin to validate parsing logic.
    
    Returns:
        dict with keys: success (bool), errors (list), warnings (list), samples (list)
    """
    from lib.ingest.plugins.dol_lca import H1BSalaryDataSourcePlugin
    from lib.ingest.plugins.dol_perm import PERMSalaryDataSourcePlugin
    from lib.ingest.registry import PluginRegistry
    from models.ingest.data_source import DataSource
    from models.ingest.enums import IngestStage, IngestStatus
    from models.ingest.ingest_run import IngestRun

    logger.info("🔬 Validating parsing for %s...", file_path.name)
    logger.info("=" * 80)

    # Register plugins
    PluginRegistry.register(H1BSalaryDataSourcePlugin(skip_clustering=True))
    PluginRegistry.register(PERMSalaryDataSourcePlugin(skip_clustering=True))

    # Detect file type from filename
    filename_lower = file_path.name.lower()
    if 'perm' in filename_lower:
        domain = 'dol'
        source_type = 'perm'
    elif 'lca' in filename_lower or 'worksite' in filename_lower:
        domain = 'dol'
        source_type = 'lca'
    else:
        return {
            'success': False,
            'errors': [f'Cannot detect file type from filename: {file_path.name}'],
            'warnings': [],
            'samples': []
        }

    plugin = PluginRegistry.get_plugin(domain, source_type)
    if not plugin:
        return {
            'success': False,
            'errors': [f'No plugin found for {domain}:{source_type}'],
            'warnings': [],
            'samples': []
        }

    logger.info("  Plugin: %s:%s", domain, source_type)
    logger.info("  Parsing %d sample rows...", sample_rows)

    # Create a mock DataSource and IngestRun for parsing (not persisted to DB)
    # The parse() method checks run.source for fiscal year extraction
    mock_source = DataSource(
        url=f'file://{file_path}',
        domain=domain,
        source_type=source_type,
        local_file_path=str(file_path)
    )

    mock_run = IngestRun(
        source=mock_source,
        status=IngestStatus.RUNNING,
        stage=IngestStage.PARSING,
        checkpoint={'filepath': str(file_path)}
    )

    errors = []
    warnings = []
    samples = []
    parsed_count = 0
    transformed_count = 0

    try:
        # Parse sample rows through plugin
        for idx, parsed_record in enumerate(plugin.parse(file_path, mock_run)):
            if idx >= sample_rows:
                break

            parsed_count += 1

            # Transform to model
            try:
                model_instance = plugin.transform(parsed_record)
                if model_instance:
                    transformed_count += 1

                    # Validate required fields
                    validation_errors = []
                    if not getattr(model_instance, 'case_number', None):
                        validation_errors.append('Missing case_number')
                    if hasattr(model_instance, 'employer_name') and not getattr(model_instance, 'employer_name', None):
                        validation_errors.append('Missing employer_name')
                    if not getattr(model_instance, 'job_title', None):
                        validation_errors.append('Missing job_title')

                    samples.append({
                        'row': idx + 1,
                        'case_number': getattr(model_instance, 'case_number', None),
                        'employer_name': getattr(model_instance, 'employer_name', None) if hasattr(model_instance, 'employer_name') else 'N/A',
                        'job_title': getattr(model_instance, 'job_title', None),
                        'validation_errors': validation_errors
                    })

                    if validation_errors:
                        errors.extend([f"Row {idx + 1}: {err}" for err in validation_errors])
                else:
                    warnings.append(f"Row {idx + 1}: transform() returned None (rejected)")
            except Exception as e:
                errors.append(f"Row {idx + 1}: Transform failed - {str(e)}")
                logger.error("Transform error on row %d: %s", idx + 1, e, exc_info=True)

    except Exception as e:
        errors.append(f"Parse failed: {str(e)}")
        logger.error("Parse error: %s", e, exc_info=True)
        return {
            'success': False,
            'errors': errors,
            'warnings': warnings,
            'samples': samples
        }

    # Log results
    logger.info("  ✓ Parsed: %d rows", parsed_count)
    logger.info("  ✓ Transformed: %d records", transformed_count)

    if samples:
        logger.info("\n  Sample records:")
        for sample in samples[:3]:  # Show first 3
            logger.info("    Row %d:", sample['row'])
            logger.info("      Case: %s", sample['case_number'])
            logger.info("      Employer: %s", sample['employer_name'][:50] if sample['employer_name'] else 'N/A')
            logger.info("      Job: %s", sample['job_title'][:50] if sample['job_title'] else 'N/A')
            if sample['validation_errors']:
                logger.info("      ⚠️  Validation errors: %s", ', '.join(sample['validation_errors']))

    if warnings:
        logger.info("\n  ⚠️  Warnings:")
        for warning in warnings[:10]:  # Show first 10
            logger.info("    - %s", warning)

    if errors:
        logger.info("\n  ❌ Errors:")
        for error in errors[:10]:  # Show first 10
            logger.info("    - %s", error)

    success = len(errors) == 0 and transformed_count > 0
    if success:
        logger.info("\n  ✅ Parsing validation PASSED")
    else:
        logger.info("\n  ❌ Parsing validation FAILED")

    return {
        'success': success,
        'errors': errors,
        'warnings': warnings,
        'samples': samples
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Inspect columns and sample values in DOL source files",
    )
    parser.add_argument(
        "--files",
        nargs="*",
        default=DEFAULT_FILES,
        help="Files to inspect (relative to workspace or absolute)",
    )
    parser.add_argument(
        "--sample-rows",
        type=int,
        default=3,
        help="Number of data rows to sample (default: 3)",
    )
    parser.add_argument(
        "--extract-row",
        type=int,
        help="Extract a specific row number (1-based, e.g., 95467 for row 95467)",
    )
    parser.add_argument(
        "--match-terms",
        type=str,
        default=",".join(DEFAULT_MATCH_TERMS),
        help="Comma-separated keywords to include (default: name,alien,beneficiary,employee,worker)",
    )
    parser.add_argument(
        "--exclude-terms",
        type=str,
        default=",".join(DEFAULT_EXCLUDE_TERMS),
        help="Comma-separated keywords to exclude (default: employer)",
    )
    parser.add_argument(
        "--show-all-columns",
        action="store_true",
        help="Include full column list in output",
    )
    parser.add_argument(
        "--estimate-storage",
        action="store_true",
        help="Estimate storage impact across inspected files",
    )
    parser.add_argument(
        "--avg-length",
        type=int,
        default=30,
        help="Average value length for storage estimate (default: 30)",
    )
    parser.add_argument(
        "--varchar-size",
        type=int,
        default=100,
        help="VARCHAR size for storage estimate (default: 100)",
    )
    parser.add_argument(
        "--validate-parsing",
        action="store_true",
        help="Smoke test: validate parsing by processing sample rows through plugin",
    )
    args = parser.parse_args()

    setup_logging(debug=False)
    logger = logging.getLogger(__name__)
    ScriptLogger(__file__).log_call(
        args={
            "files": args.files,
            "sample_rows": args.sample_rows,
            "match_terms": args.match_terms,
            "exclude_terms": args.exclude_terms,
            "show_all_columns": args.show_all_columns,
            "estimate_storage": args.estimate_storage,
            "avg_length": args.avg_length,
            "varchar_size": args.varchar_size,
        },
        context="Inspecting source file columns",
    )

    files = _resolve_files(args.files)
    match_terms = [term.strip() for term in args.match_terms.split(",") if term.strip()]
    exclude_terms = [term.strip() for term in args.exclude_terms.split(",") if term.strip()]

    # Validation mode: smoke test parsing
    if args.validate_parsing:
        logger.info("🔬 SMOKE TEST: Validating parsing for %d files", len(files))
        logger.info("=" * 80)

        validation_results = []
        for file_path in files:
            if not file_path.exists():
                logger.warning("%s not found", file_path)
                continue

            result = _validate_parsing(logger, file_path, sample_rows=5)
            validation_results.append((file_path.name, result))
            logger.info("")

        # Summary
        logger.info("=" * 80)
        logger.info("SMOKE TEST SUMMARY")
        logger.info("=" * 80)
        passed = sum(1 for _, r in validation_results if r['success'])
        failed = len(validation_results) - passed
        logger.info("Total files: %d", len(validation_results))
        logger.info("  ✅ Passed: %d", passed)
        logger.info("  ❌ Failed: %d", failed)

        if failed > 0:
            logger.info("\nFailed files:")
            for filename, result in validation_results:
                if not result['success']:
                    logger.info("  - %s", filename)
                    for error in result['errors'][:3]:
                        logger.info("      %s", error)
            sys.exit(1)  # Exit with error if any validation failed
        else:
            logger.info("\n✅ All files passed smoke test - ready for ingestion!")

        return

    # Extract specific row mode
    if args.extract_row:
        target_row = args.extract_row
        logger.info("Extracting row %d from %d files", target_row, len(files))
        logger.info("=" * 80)

        for file_path in files:
            if not file_path.exists():
                logger.warning("%s not found", file_path)
                continue

            logger.info("\nFile: %s", file_path.name)
            logger.info("-" * 80)

            # Read header and target row
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Get header
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            logger.info("Header (%d columns):", len(header_row))

            # Get target row
            found = False
            for i, row in enumerate(ws.iter_rows(min_row=target_row, max_row=target_row, values_only=True), start=target_row):
                found = True
                logger.info("\nRow %d values:", i)

                # Show all non-None values
                for col_name, value in zip(header_row, row):
                    if value is not None:
                        logger.info("  %s: %r", col_name, value)

                # Focus on wage fields
                logger.info("\n🔍 Key wage fields:")
                wage_cols = ['CASE_NUMBER', 'WAGE_RATE_OF_PAY_FROM', 'WAGE_RATE_OF_PAY_TO', 'WAGE_UNIT_OF_PAY', 'PW_UNIT_OF_PAY']
                for col in wage_cols:
                    try:
                        idx = list(header_row).index(col)
                        logger.info("  %s: %r", col, row[idx])
                    except (ValueError, IndexError):
                        pass

            wb.close()

            if not found:
                logger.warning("Row %d not found in %s", target_row, file_path.name)

        return

    # Normal inspection mode
    total_records = 0
    for file_path in files:
        if not file_path.exists():
            logger.warning("%s not found", file_path)
            continue

        columns = read_excel_headers(file_path)
        _log_column_samples(
            logger,
            file_path,
            columns,
            args.sample_rows,
            match_terms,
            exclude_terms,
            args.show_all_columns,
        )

        file_stats = get_file_stats(file_path, logger_instance=logger)
        total_records += file_stats["row_count"]
        logger.info("Total records in %s: %s", file_path.name, f"{file_stats['row_count']:,}")
        logger.info("")

    if args.estimate_storage and total_records:
        _log_storage_estimate(logger, total_records, args.avg_length, args.varchar_size)


if __name__ == "__main__":
    main()
