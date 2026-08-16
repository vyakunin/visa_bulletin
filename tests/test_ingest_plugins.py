"""Unit tests for ingest plugins"""

# Use shared Django setup
from tests.django_setup import setup_django_for_tests

setup_django_for_tests()

from pathlib import Path
from unittest.mock import Mock, patch

import pytest

from lib.ingest.plugins.dol_lca import H1BSalaryDataSourcePlugin
from lib.ingest.plugins.dol_perm import PERMSalaryDataSourcePlugin
from lib.ingest.registry import PluginRegistry
from models.enums.visa_program import VisaProgram
from models.ingest.data_source import DataSource
from models.ingest.enums import DataDomain, FormatVersion, IngestStatus, SourceType
from models.ingest.ingest_run import IngestRun
from models.salary import SalaryRecord, WorksiteRecord


@pytest.mark.django_db
class TestH1BSalaryDataSourcePlugin:
    """Tests for H-1B LCA salary data source plugin"""

    def test_plugin_attributes(self):
        """Test plugin has correct domain and source type"""
        plugin = H1BSalaryDataSourcePlugin()

        assert plugin.domain == DataDomain.DOL
        assert plugin.source_type == SourceType.LCA

    def test_get_format_version(self):
        """Test format version detection from filename"""
        plugin = H1BSalaryDataSourcePlugin()

        # FY >= 2015 → MODERN format (get_format_version returns a FormatVersion enum).
        filepath = Path("LCA_Disclosure_Data_FY2024.xlsx")
        version = plugin.get_format_version(filepath)
        assert version == FormatVersion.MODERN

        # Test with quarter
        filepath = Path("LCA_Disclosure_Data_FY2024_Q4.xlsx")
        version = plugin.get_format_version(filepath)
        assert version == FormatVersion.MODERN

    @patch("lib.ingest.plugins.dol_lca.fetch_page")
    def test_discover_sources(self, mock_fetch):
        """Test source discovery"""
        plugin = H1BSalaryDataSourcePlugin()

        # Mock HTML response with LCA links
        mock_fetch.return_value = """
        <html>
        <a href="LCA_Disclosure_Data_FY2024.xlsx">Download</a>
        <a href="LCA_Disclosure_Data_FY2023.xlsx">Download</a>
        </html>
        """

        sources = plugin.discover_sources()

        assert len(sources) >= 2
        assert all(s.domain == DataDomain.DOL.value for s in sources)
        assert all(s.source_type == SourceType.LCA.value for s in sources)

    # download() is inherited from the base plugin. base.download() builds its own
    # dest_path under get_workspace_dir()/data/..., calls download_file(url, dest_path)
    # (return value is ignored), then compute_file_hash(dest_path) — so the mock must
    # materialize dest_path or the hash step raises FileNotFoundError. Both helpers are
    # imported locally inside download() from lib.utils.http_utils, so patch them there.
    # A tmp workspace keeps dest_path out of the read-only runfiles tree and guarantees
    # it doesn't pre-exist between runs (else the file-exists early-return path skips the
    # download_file call entirely).
    @patch("lib.utils.http_utils.get_workspace_dir")
    @patch("lib.utils.http_utils.download_file")
    def test_download(self, mock_download, mock_workspace, tmp_path):
        """download() materializes the file, hashes it, persists metadata, returns dest_path."""
        mock_workspace.return_value = tmp_path

        plugin = H1BSalaryDataSourcePlugin()

        source = DataSource.objects.create(
            url="https://example.com/data.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )

        run = IngestRun.objects.create(source=source, status=IngestStatus.PENDING)

        def fake_download(url, dest_path):
            Path(dest_path).write_bytes(b"test-content")

        mock_download.side_effect = fake_download

        result = plugin.download(source, run)

        mock_download.assert_called_once()
        called_url, called_dest = mock_download.call_args[0]
        assert called_url == source.url
        # download() returns the dest_path it built, not download_file's return value.
        assert result == Path(called_dest)
        assert result.exists()

        # Metadata persisted from the real hash of the materialized file.
        source.refresh_from_db()
        assert source.content_hash
        assert source.local_file_path == str(result)

    def test_parse_excel_streaming(self, tmp_path):
        """Test Excel parsing with openpyxl streaming"""
        plugin = H1BSalaryDataSourcePlugin()

        # Create a simple Excel file for testing
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME", "JOB_TITLE"])
        ws.append(["CASE1", "Company A", "Engineer"])
        ws.append(["CASE2", "Company B", "Manager"])

        test_file = tmp_path / "test.xlsx"
        wb.save(test_file)

        source = DataSource.objects.create(
            url="https://example.com/test.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )

        run = IngestRun.objects.create(
            source=source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 2
        assert records[0]["CASE_NUMBER"] == "CASE1"
        assert records[0]["EMPLOYER_NAME"] == "Company A"
        assert records[1]["CASE_NUMBER"] == "CASE2"

    def test_fiscal_year_extraction_4digit(self, tmp_path):
        """Test fiscal year extraction from 4-digit filename during parse"""
        plugin = H1BSalaryDataSourcePlugin()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME"])
        ws.append(["CASE1", "Company A"])

        test_file = tmp_path / "H-1B_Disclosure_Data_FY2018_EOY.xlsx"
        wb.save(test_file)

        source = DataSource.objects.create(
            url="https://example.com/H-1B_Disclosure_Data_FY2018_EOY.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )

        run = IngestRun.objects.create(
            source=source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 1
        assert records[0]["_fiscal_year"] == 2018
        assert records[0]["_source_file"] == "H-1B_Disclosure_Data_FY2018_EOY.xlsx"

    def test_fiscal_year_extraction_2digit(self, tmp_path):
        """Test fiscal year extraction from 2-digit filename (FY17, FY16, FY14) during parse"""
        plugin = H1BSalaryDataSourcePlugin()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME"])
        ws.append(["CASE1", "Company A"])

        # Test FY17 -> 2017
        test_file = tmp_path / "H-1B_Disclosure_Data_FY17.xlsx"
        wb.save(test_file)

        source = DataSource.objects.create(
            url="https://example.com/H-1B_Disclosure_Data_FY17.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )

        run = IngestRun.objects.create(
            source=source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 1
        assert records[0]["_fiscal_year"] == 2017

        # Test FY14_Q4 -> 2014
        test_file2 = tmp_path / "H-1B_FY14_Q4.xlsx"
        wb.save(test_file2)

        source2 = DataSource.objects.create(
            url="https://example.com/H-1B_FY14_Q4.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )

        run2 = IngestRun.objects.create(
            source=source2, status=IngestStatus.PENDING, checkpoint={}
        )

        records2 = list(plugin.parse(test_file2, run2))

        assert len(records2) == 1
        assert records2[0]["_fiscal_year"] == 2014

    def test_fiscal_year_extraction_url_fallback_artificial_filename(self, tmp_path):
        """Test fiscal year extraction with URL fallback for artificial filenames (lca_xxx.xlsx)"""
        plugin = H1BSalaryDataSourcePlugin()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME"])
        ws.append(["CASE1", "Company A"])

        # Artificial filename (no fiscal year in filename)
        test_file = tmp_path / "lca_362.xlsx"
        wb.save(test_file)

        # DataSource URL has original filename with fiscal year
        source = DataSource.objects.create(
            url="file://H-1B_Disclosure_Data_FY17.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )

        run = IngestRun.objects.create(
            source=source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 1
        # Should extract fiscal year from URL fallback (FY17 -> 2017)
        assert records[0]["_fiscal_year"] == 2017
        assert records[0]["_source_file"] == "lca_362.xlsx"

    def test_fiscal_year_extraction_url_fallback_2digit(self, tmp_path):
        """Test fiscal year extraction with URL fallback for 2-digit years in URL"""
        plugin = H1BSalaryDataSourcePlugin()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME"])
        ws.append(["CASE1", "Company A"])

        # Artificial filename
        test_file = tmp_path / "lca_365.xlsx"
        wb.save(test_file)

        # DataSource URL has 2-digit fiscal year (FY14 -> 2014)
        source = DataSource.objects.create(
            url="file://H-1B_FY14_Q4.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )

        run = IngestRun.objects.create(
            source=source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 1
        # Should extract fiscal year from URL fallback (FY14 -> 2014)
        assert records[0]["_fiscal_year"] == 2014

    def test_fiscal_year_extraction_alternative_datasource(self, tmp_path):
        """Test fiscal year extraction using alternative DataSource (sophisticated fallback)"""
        plugin = H1BSalaryDataSourcePlugin()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME"])
        ws.append(["CASE1", "Company A"])

        # Artificial filename (like lca_368.xlsx scenario)
        test_file = tmp_path / "lca_368.xlsx"
        wb.save(test_file)
        file_path_str = str(test_file.absolute())

        # Create reimport:// DataSource (no fiscal year in URL)
        reimport_source = DataSource.objects.create(
            url=f"reimport://{file_path_str}",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
            local_file_path=file_path_str,
        )

        # Create alternative DataSource with same local_file_path but original filename in URL
        _alt_source = DataSource.objects.create(
            url="file://H-1B_Disclosure_Data_FY15_Q4.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
            local_file_path=file_path_str,  # Same file!
        )

        run = IngestRun.objects.create(
            source=reimport_source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 1
        # Should extract fiscal year from alternative DataSource (FY15 -> 2015)
        assert records[0]["_fiscal_year"] == 2015


@pytest.mark.django_db
class TestPluginRegistry:
    """Tests for plugin registry"""

    def test_register_plugin(self):
        """Test plugin registration"""
        PluginRegistry.clear()

        plugin = H1BSalaryDataSourcePlugin()
        PluginRegistry.register(plugin)

        retrieved = PluginRegistry.get_plugin(DataDomain.DOL, SourceType.LCA)
        assert retrieved == plugin

    def test_get_plugin_by_string(self):
        """Test getting plugin by string values"""
        PluginRegistry.clear()

        plugin = H1BSalaryDataSourcePlugin()
        PluginRegistry.register(plugin)

        retrieved = PluginRegistry.get_plugin(
            DataDomain.DOL.value, SourceType.LCA.value
        )
        assert retrieved == plugin

    def test_get_nonexistent_plugin(self):
        """Test getting non-existent plugin returns None"""
        PluginRegistry.clear()

        plugin = PluginRegistry.get_plugin(
            DataDomain.VISA_BULLETIN, SourceType.BULLETIN
        )
        assert plugin is None

    def test_list_plugins(self):
        """Test listing all plugins"""
        PluginRegistry.clear()

        plugin1 = H1BSalaryDataSourcePlugin()
        PluginRegistry.register(plugin1)

        plugins = PluginRegistry.list_plugins()
        assert len(plugins) == 1
        assert plugins[0][0] == DataDomain.DOL.value
        assert plugins[0][1] == SourceType.LCA.value


@pytest.mark.django_db
class TestPERMSalaryDataSourcePlugin:
    """Tests for PERM salary data source plugin"""

    def test_plugin_attributes(self):
        """Test plugin has correct domain and source type"""
        plugin = PERMSalaryDataSourcePlugin()

        assert plugin.domain == DataDomain.DOL
        assert plugin.source_type == SourceType.PERM

    def test_transform_skips_record_without_employer_name(self):
        """Test transform skips records without employer_name"""
        plugin = PERMSalaryDataSourcePlugin()
        plugin._current_run = Mock()

        # Record with case number but no employer name
        record = {
            "CASE_NUMBER": "P-12345-67890",
            "JOB_TITLE": "Engineer",
            "WAGE_OFFER_FROM_9089": "120000",
            "WAGE_OFFER_UNIT_OF_PAY_9089": "Year",
            "_fiscal_year": 2024,
            "_source_file": "PERM_FY2024.xlsx",
            # Missing EMPLOYER_NAME
        }

        result = plugin.transform(record)
        assert result is None

    def test_transform_skips_record_without_salary_data(self):
        """Test transform skips records without salary data"""
        plugin = PERMSalaryDataSourcePlugin()
        plugin._current_run = Mock()

        # Record with employer but no salary data
        record = {
            "CASE_NUMBER": "P-12345-67890",
            "EMPLOYER_NAME": "Tech Corp",
            "EMPLOYER_CITY": "Seattle",
            "EMPLOYER_STATE": "WA",
            "JOB_TITLE": "Engineer",
            "_fiscal_year": 2024,
            "_source_file": "PERM_FY2024.xlsx",
            # Missing WAGE_OFFER_FROM_9089 and WAGE_OFFER_UNIT_OF_PAY_9089
        }

        result = plugin.transform(record)
        assert result is None

    def test_transform_creates_record_with_all_required_fields(self):
        """Test transform creates record when all required fields are present"""
        plugin = PERMSalaryDataSourcePlugin()
        plugin._current_run = Mock()

        # Record with all required fields
        record = {
            "CASE_NUMBER": "P-12345-67890",
            "EMPLOYER_NAME": "Tech Corp",
            "EMPLOYER_CITY": "Seattle",
            "EMPLOYER_STATE": "WA",
            "JOB_TITLE": "Engineer",
            "WAGE_OFFER_FROM_9089": "120000",
            "WAGE_OFFER_UNIT_OF_PAY_9089": "Year",
            "_fiscal_year": 2024,
            "_source_file": "PERM_FY2024.xlsx",
        }

        result = plugin.transform(record)
        assert result is not None
        assert isinstance(result, SalaryRecord)
        assert result.case_number == "P-12345-67890"
        assert result.employer_name == "Tech Corp"
        assert result.wage_annual == 120000.0

    def test_fiscal_year_extraction_2digit_perm(self, tmp_path):
        """Test fiscal year extraction from 2-digit filename for PERM plugin"""
        plugin = PERMSalaryDataSourcePlugin()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME"])
        ws.append(["P-12345", "Company A"])

        # Test FY16 -> 2016
        test_file = tmp_path / "PERM_FY16.xlsx"
        wb.save(test_file)

        source = DataSource.objects.create(
            url="https://example.com/PERM_FY16.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.PERM,
        )

        run = IngestRun.objects.create(
            source=source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 1
        assert records[0]["_fiscal_year"] == 2016

    def test_fiscal_year_extraction_url_fallback_perm(self, tmp_path):
        """Test fiscal year extraction with URL fallback for PERM artificial filenames"""
        plugin = PERMSalaryDataSourcePlugin()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["CASE_NUMBER", "EMPLOYER_NAME"])
        ws.append(["P-12345", "Company A"])

        # Artificial filename
        test_file = tmp_path / "perm_123.xlsx"
        wb.save(test_file)

        # DataSource URL has original filename with 2-digit fiscal year
        source = DataSource.objects.create(
            url="file://PERM_FY16.xlsx",
            domain=DataDomain.DOL,
            source_type=SourceType.PERM,
        )

        run = IngestRun.objects.create(
            source=source, status=IngestStatus.PENDING, checkpoint={}
        )

        records = list(plugin.parse(test_file, run))

        assert len(records) == 1
        # Should extract fiscal year from URL fallback (FY16 -> 2016)
        assert records[0]["_fiscal_year"] == 2016


@pytest.mark.django_db
class TestWorksiteFileTransformFy2026:
    """Regression: FY2026 standalone worksite files (pw_worksites/lca_worksites)
    carry only CASE_NUMBER + worksite location + SOC columns and have NO
    JOB_TITLE. The transform required job_title and dropped every row -> 0
    records ingested. It must now create a WorksiteRecord, falling back to the
    SOC title for the missing job_title.
    """

    def _worksite_file_row(self):
        # Column subset taken verbatim from pw_worksites_fy2026_q2.xlsx.
        return {
            "CASE_NUMBER": "I-200-26042-700001",
            "WORKSITE_COUNTY": "SANTA CLARA",
            "WORKSITE_STATE": "CA",
            "WORKSITE_BLS_AREA": "41940",
            "SOC_CODE": "15-1252",
            "SOC_TITLE": "Software Developers",
            "_fiscal_year": 2026,
            "_source_file": "pw_worksites_fy2026_q2.xlsx",
        }

    def test_worksite_file_row_creates_record_without_job_title(self):
        plugin = H1BSalaryDataSourcePlugin()
        plugin._current_run = Mock()
        result = plugin.transform(self._worksite_file_row())
        assert isinstance(result, WorksiteRecord), (
            "worksite-file row (no JOB_TITLE) must yield a WorksiteRecord, not be dropped"
        )
        assert result.worksite_state == "CA"
        assert result.soc_code == "15-1252"
        # job_title falls back to SOC title when the file omits JOB_TITLE.
        assert result.job_title == "Software Developers"


@pytest.mark.django_db
class TestWorksiteFileValidationFy2026:
    """Regression: a `pw_worksites_*` file's rows carry PREVAILING-WAGE case
    numbers (`P-...`), which transform() maps to VisaProgram.PERM. The
    post-ingest validator looked the rows up filtered on VisaProgram.H1B, found
    none, and failed the run with "expected data but got none" while the rows
    sat committed in the table. That made every quarter's DOL ingest exit 1.

    The worksite lookup is program-agnostic now: a worksite file's program is a
    property of its case numbers, not of the plugin that read it.
    """

    SOURCE_FILE = "pw_worksites_fy2026_q3.xlsx"

    def _run_for(self, filename):
        source = DataSource.objects.create(
            url=f"https://www.dol.gov/sites/dolgov/files/eta/oflc/pdfs/fy26q3/{filename}",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
        )
        return IngestRun.objects.create(
            source=source,
            status=IngestStatus.RUNNING,
            checkpoint={"filepath": f"/tmp/{filename}"},
        )

    def _worksite_row(self, case_number, source_file):
        # Shaped like a real pw_worksites row: no employer, no wage, SOC title
        # standing in for the absent JOB_TITLE column.
        return WorksiteRecord.objects.create(
            case_number=case_number,
            visa_program=VisaProgram.PERM,
            worksite_city="SANTA CLARA",
            worksite_state="CA",
            job_title="Software Developers",
            soc_code="15-1252",
            soc_title="Software Developers",
            fiscal_year=2026,
            source_file=source_file,
        )

    def test_prevailing_wage_worksite_rows_validate_clean(self):
        """The defect: PERM-programmed worksite rows exist, so the run must pass."""
        run = self._run_for(self.SOURCE_FILE)
        for n in range(3):
            self._worksite_row(f"P-100-25024-64097{n}", self.SOURCE_FILE)

        result = H1BSalaryDataSourcePlugin().validate_post_ingest(run)

        assert result.passed, (
            "worksite rows are committed under VisaProgram.PERM (P- case numbers); "
            f"the validator must count them, not fail the run: {result.errors}"
        )
        assert not any("expected data but got none" in e for e in result.errors)

    def test_source_file_that_created_nothing_still_fails(self):
        """The boundary: the fix must not turn the validator into a rubber stamp.

        Same code path, no rows -> the run must still fail. Without this, a
        program-agnostic lookup could pass any run that touched a table with
        rows in it from some other file.
        """
        run = self._run_for("pw_worksites_fy2026_q4.xlsx")
        # Rows exist, but from a DIFFERENT source file — they must not be counted.
        self._worksite_row("P-100-25024-999999", self.SOURCE_FILE)

        result = H1BSalaryDataSourcePlugin().validate_post_ingest(run)

        assert not result.passed, "a source file that created no rows must fail"
        assert any("expected data but got none" in e for e in result.errors), (
            f"expected the no-records error, got: {result.errors}"
        )
