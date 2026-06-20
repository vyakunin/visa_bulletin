"""Integration tests for ingest pipeline"""

# Use shared Django setup
from tests.django_setup import setup_django_for_tests

setup_django_for_tests()

from unittest.mock import MagicMock, Mock, patch

import pytest
from django.db import connection
from openpyxl import Workbook

from lib.ingest.base import ValidationResult
from lib.ingest.orchestrator import PipelineOrchestrator
from lib.ingest.plugins.dol_lca import H1BSalaryDataSourcePlugin
from lib.ingest.plugins.dol_perm import PERMSalaryDataSourcePlugin
from lib.ingest.plugins.visa_bulletin import VisaBulletinPlugin
from lib.ingest.registry import PluginRegistry
from models.ingest.data_source import DataSource
from models.ingest.enums import (
    DataDomain,
    FormatVersion,
    IngestStage,
    IngestStatus,
    SourceType,
)
from models.salary import SalaryRecord


@pytest.mark.django_db
class TestIngestPipelineIntegration:
    """Integration tests for full pipeline"""

    def test_pipeline_handles_missing_record_estimate(self, tmp_path):
        """Pipeline should not crash if row count cannot be estimated."""
        source = DataSource.objects.create(
            url="file://missing.html",
            domain=DataDomain.VISA_BULLETIN,
            source_type=SourceType.BULLETIN,
            format_version=FormatVersion.MODERN,
        )

        plugin = MagicMock()
        plugin.set_rejection_tracker = Mock()

        orchestrator = PipelineOrchestrator()
        orchestrator.update_mode = True  # Skip versioning logic for unit-style test

        dummy_file = tmp_path / "dummy.html"
        dummy_file.write_text("test")

        with (
            patch(
                "lib.ingest.orchestrator.PluginRegistry.get_plugin", return_value=plugin
            ),
            patch(
                "lib.ingest.orchestrator.get_data_source_filepath", return_value=None
            ),
            patch("lib.ingest.orchestrator.RejectionTracker") as rejection_tracker,
            patch.object(orchestrator, "_download_stage", return_value=dummy_file),
            patch.object(orchestrator, "_parse_stage", return_value=[]),
            patch.object(orchestrator, "_transform_stage", return_value=[]),
            patch.object(orchestrator, "_load_to_db_stage", return_value=None),
            patch.object(
                orchestrator,
                "_validate_post_ingest",
                return_value=ValidationResult(passed=True),
            ),
        ):
            rejection_tracker.return_value.save_to_db = Mock()
            run = orchestrator.run(source, resume=False)

        assert run.status == IngestStatus.COMPLETED

    def test_full_pipeline_with_excel(self, tmp_path):
        """Test full pipeline: discover -> download -> parse -> transform -> load"""
        # Register plugins
        PluginRegistry.register(H1BSalaryDataSourcePlugin())
        PluginRegistry.register(PERMSalaryDataSourcePlugin())
        PluginRegistry.register(VisaBulletinPlugin())

        # Create test Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "CASE_NUMBER",
                "EMPLOYER_NAME",
                "JOB_TITLE",
                "WAGE_RATE_OF_PAY_FROM",
                "WAGE_UNIT_OF_PAY",
            ]
        )
        ws.append(["CASE001", "Test Company", "Software Engineer", "150000", "Year"])
        ws.append(["CASE002", "Test Company", "Data Scientist", "140000", "Year"])

        test_file = tmp_path / "test_lca_FY2024.xlsx"
        wb.save(test_file)

        # Create data source
        source = DataSource.objects.create(
            url=f"file://{test_file}",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
            format_version=FormatVersion.MODERN,
        )

        # Mock download to return test file
        plugin = PluginRegistry.get_plugin(DataDomain.DOL, SourceType.LCA)
        original_download = plugin.download

        def mock_download(s, r):
            return test_file

        plugin.download = mock_download

        # Run pipeline
        orchestrator = PipelineOrchestrator(
            batch_size=10, adaptive_batch=False, prefilter_existing=False
        )

        run = orchestrator.run(source, resume=False)

        # Verify run completed
        assert run.status == IngestStatus.COMPLETED
        assert run.stage == IngestStage.COMPLETED
        assert run.records_created == 2

        # Verify records in database
        records = SalaryRecord.objects.filter(source_file=test_file.name)
        assert records.count() == 2

        # Verify data
        record1 = records.get(case_number="CASE001")
        assert record1.employer_name == "Test Company"
        assert record1.job_title == "Software Engineer"
        assert float(record1.wage_annual) == 150000.0

        # Restore original method
        plugin.download = original_download

    def test_rerun_is_idempotent_no_duplicates(self, tmp_path):
        """Re-running ingest on the same source converges to the same rows, never doubles.

        This is the practical 'resume' guarantee. ``orchestrator.run(resume=True)``
        does NOT continue a prior RUNNING run in place — ``_get_or_create_run`` skips
        RUNNING runs ("another process is running") and creates a fresh run, which
        re-processes the whole file and dedups by case_number. So the meaningful
        invariant is idempotency: after a second run the DB still holds exactly the
        original row count and the second run creates 0 new records.

        (The checkpoint-skip path in ``_parse_stage`` is vestigial for this entry
        point; resilience comes from dedup, not from resuming the same run object.)
        """
        # Register plugins
        PluginRegistry.register(H1BSalaryDataSourcePlugin())
        PluginRegistry.register(PERMSalaryDataSourcePlugin())
        PluginRegistry.register(VisaBulletinPlugin())

        # Create test Excel file with many rows
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "CASE_NUMBER",
                "EMPLOYER_NAME",
                "JOB_TITLE",
                "WAGE_RATE_OF_PAY_FROM",
                "WAGE_UNIT_OF_PAY",
            ]
        )
        for i in range(50):
            ws.append([f"CASE{i:03d}", "Test Company", "Engineer", "100000", "Year"])

        test_file = tmp_path / "test_lca_large_FY2024.xlsx"
        wb.save(test_file)

        source = DataSource.objects.create(
            url=f"file://{test_file}", domain=DataDomain.DOL, source_type=SourceType.LCA
        )

        plugin = PluginRegistry.get_plugin(DataDomain.DOL, SourceType.LCA)
        original_download = plugin.download

        def mock_download(s, r):
            return test_file

        plugin.download = mock_download

        try:
            orchestrator = PipelineOrchestrator(batch_size=10, adaptive_batch=False)

            # First run: ingests all 50 rows.
            run1 = orchestrator.run(source, resume=False)
            assert run1.status == IngestStatus.COMPLETED
            assert run1.records_created == 50
            assert (
                SalaryRecord.objects.filter(source_file=test_file.name).count() == 50
            )

            # Second run (resume=True): must converge to the same 50, not duplicate.
            run2 = orchestrator.run(source, resume=True)
            assert run2.status == IngestStatus.COMPLETED
            assert (
                SalaryRecord.objects.filter(source_file=test_file.name).count() == 50
            ), "re-run must not create duplicate records"
            assert run2.records_created == 0, (
                "every row already present -> second run dedups, creates nothing"
            )
        finally:
            plugin.download = original_download

    def test_plugin_registration(self):
        """Test that plugins are registered correctly"""
        PluginRegistry.clear()
        # Register plugins
        PluginRegistry.register(H1BSalaryDataSourcePlugin())
        PluginRegistry.register(PERMSalaryDataSourcePlugin())
        PluginRegistry.register(VisaBulletinPlugin())

        # Check DOL LCA plugin
        lca_plugin = PluginRegistry.get_plugin(DataDomain.DOL, SourceType.LCA)
        assert lca_plugin is not None
        assert lca_plugin.domain == DataDomain.DOL
        assert lca_plugin.source_type == SourceType.LCA

        # Check DOL PERM plugin
        perm_plugin = PluginRegistry.get_plugin(DataDomain.DOL, SourceType.PERM)
        assert perm_plugin is not None

        # Check list
        plugins = PluginRegistry.list_plugins()
        assert len(plugins) >= 2

    @pytest.mark.django_db(transaction=True)
    def test_db_error_retry(self, tmp_path):
        """
        A DB error during load fails the run and persists nothing; a rerun after
        the DB is fixed succeeds and saves the records.

        Scenario:
        1. Pipeline runs but the target table is missing (database error)
        2. Run is marked FAILED with records_failed > 0; nothing persisted
        3. The table is restored (fixing the issue)
        4. Pipeline is rerun
        5. Records are successfully saved on rerun

        Uses the salary (LCA Excel) path: it parses cleanly and routes the missing
        table through the load-batch error handler (so records_failed is counted),
        unlike the bulletin-HTML path whose toy fixture no longer parses into any
        records.

        Requires ``transaction=True``: the test deliberately drops a table,
        producing a real ``relation ... does not exist`` error. Under the default
        (transaction-wrapped) django_db that error aborts the single outer
        transaction so every subsequent query — including the orchestrator's own
        ``mark_failed`` — raises ``TransactionManagementError``. With real
        per-statement transactions the orchestrator recovers exactly as in prod.
        """
        # Register plugins
        PluginRegistry.register(H1BSalaryDataSourcePlugin())
        PluginRegistry.register(PERMSalaryDataSourcePlugin())
        PluginRegistry.register(VisaBulletinPlugin())

        # Create test LCA Excel file (FY in the name so fiscal_year is derived)
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "CASE_NUMBER",
                "EMPLOYER_NAME",
                "JOB_TITLE",
                "WAGE_RATE_OF_PAY_FROM",
                "WAGE_UNIT_OF_PAY",
            ]
        )
        ws.append(["CASE001", "Test Company", "Software Engineer", "150000", "Year"])
        ws.append(["CASE002", "Test Company", "Data Scientist", "140000", "Year"])

        test_file = tmp_path / "test_lca_FY2024.xlsx"
        wb.save(test_file)

        source = DataSource.objects.create(
            url=f"file://{test_file}",
            domain=DataDomain.DOL,
            source_type=SourceType.LCA,
            format_version=FormatVersion.MODERN,
        )

        plugin = PluginRegistry.get_plugin(DataDomain.DOL, SourceType.LCA)
        original_download = plugin.download

        def mock_download(s, r):
            return test_file

        plugin.download = mock_download

        # Temporarily rename salary_record table to simulate a missing-table error
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT 1 FROM information_schema.tables WHERE table_schema = 'public' AND table_name = 'salary_record'"
            )
            if cursor.fetchone():
                cursor.execute(
                    "ALTER TABLE salary_record RENAME TO salary_record_backup"
                )

        # Run everything under try/finally so a renamed table is ALWAYS restored
        # even if an assertion fails mid-test — a left-renamed `salary_record`
        # would break every subsequent transaction=True test in this module.
        try:
            orchestrator = PipelineOrchestrator(
                batch_size=10, adaptive_batch=False, prefilter_existing=False
            )

            # First run - should fail with database error
            try:
                run = orchestrator.run(source, resume=False)
                # Should not reach here, but if it does, verify it failed
                assert run.status == IngestStatus.FAILED, (
                    "Expected run to fail with missing table"
                )
            except Exception as e:
                # Expected - PostgreSQL missing-table error ("does not exist" / "relation")
                err = str(e).lower()
                assert (
                    "no such table" in err
                    or "does not exist" in err
                    or "relation" in err
                    or "salary_record" in err
                ), f"Expected DB error, got: {e}"

            # The run that was created should be marked FAILED with failures counted.
            run = source.runs.order_by("-started_at").first()
            assert run is not None
            assert run.status == IngestStatus.FAILED
            assert run.records_failed > 0, (
                "Should have records_failed > 0 when the load batch fails"
            )

            # Restore table (fix the database issue) BEFORE asserting counts —
            # `SalaryRecord.objects.count()` would itself raise while the table is
            # renamed away, so the "nothing persisted" check must run after restore.
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT 1 FROM information_schema.tables WHERE table_schema = 'public' AND table_name = 'salary_record_backup'"
                )
                if cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE salary_record_backup RENAME TO salary_record"
                    )

            # Verify the failed run persisted nothing (it committed no rows).
            assert (
                SalaryRecord.objects.filter(source_file=test_file.name).count() == 0
            ), "No salary records should be saved on failed run"

            # Rerun pipeline - should succeed now
            run_retry = orchestrator.run(source, resume=True)

            # Verify run completed successfully
            assert run_retry.status == IngestStatus.COMPLETED, (
                f"Run should complete on retry, got status: {run_retry.status}"
            )
            assert run_retry.stage == IngestStage.COMPLETED
            assert run_retry.records_created == 2, (
                "Should have created both records on successful retry"
            )
            assert run_retry.records_failed == 0, (
                "Should have no failed records on successful retry"
            )

            # Verify records were actually saved to database
            assert (
                SalaryRecord.objects.filter(source_file=test_file.name).count() == 2
            ), "Salary records should be saved on retry"
        finally:
            # Always restore: rename the table back if still renamed, restore download.
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT 1 FROM information_schema.tables WHERE table_schema = 'public' AND table_name = 'salary_record_backup'"
                )
                if cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE salary_record_backup RENAME TO salary_record"
                    )
            plugin.download = original_download
